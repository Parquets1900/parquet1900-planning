#!/usr/bin/env python3
"""
Parquet 1900 — Generador de datos para la app de planning
Uso: python3 generar_datos.py ruta/al/archivo.xlsx
Genera data.json en la misma carpeta que este script.
"""

import sys
import json
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles.colors import Color
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

MESES = ["junio", "julio", "agosto", "septiembre"]
MESES_ES = ["Junio", "Julio", "Agosto", "Septiembre"]

def color_rgb(cell):
    """Devuelve el color de fondo de la celda como string 'RRGGBB' o None."""
    try:
        fill = cell.fill
        if fill is None:
            return None
        fg = fill.fgColor
        if fg is None:
            return None
        # Color directo en RGB/ARGB
        if fg.type == "rgb":
            argb = fg.rgb  # e.g. "FFFFFF00"
            if len(argb) == 8:
                return argb[2:]  # quitar alfa
            if len(argb) == 6:
                return argb
        # Color indexado — tabla estándar de Excel (simplificada)
        if fg.type == "indexed":
            idx_map = {
                3:  "FF0000",  # rojo
                4:  "00FF00",  # verde
                5:  "0000FF",  # azul
                6:  "FFFF00",  # amarillo
                7:  "FF00FF",
                27: "FFFF00",  # amarillo variante
                10: "FF0000",
                13: "FFFF00",
            }
            return idx_map.get(fg.indexed)
        # Color de tema — aproximación básica
        if fg.type == "theme":
            return None  # difícil sin resolver la paleta del tema
    except Exception:
        pass
    return None

def is_yellow(rgb):
    if not rgb or len(rgb) < 6:
        return False
    r, g, b = int(rgb[0:2],16), int(rgb[2:4],16), int(rgb[4:6],16)
    return r > 160 and g > 160 and b < 80

def is_red(rgb):
    if not rgb or len(rgb) < 6:
        return False
    r, g, b = int(rgb[0:2],16), int(rgb[2:4],16), int(rgb[4:6],16)
    return r > 160 and g < 80 and b < 80

def parse_excel(filepath):
    print(f"Abriendo: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    print(f"Hoja activa: {ws.title}  |  Filas: {ws.max_row}  Columnas: {ws.max_column}")

    result = {m: {} for m in MESES_ES}

    cur_mes = None
    day_col_map = {}   # col_index → day_number
    fixed_cols  = {}   # 'nprsto'|'nfra'|'cliente'|'obra'|'nobre' → col_index

    rows = list(ws.iter_rows())

    for ri, row in enumerate(rows):
        # Valores de todas las celdas de la fila (cadena limpia)
        vals = [str(cell.value or "").strip() for cell in row]

        # ── ¿Es fila de cabecera? Buscar NPRSTO en cualquier celda
        nprsto_ci = None
        for ci, v in enumerate(vals):
            if "PREST" in v.upper() or v.upper() == "NPRSTO":
                nprsto_ci = ci
                break

        if nprsto_ci is not None:
            # Nueva sección — reiniciar mapeo
            day_col_map = {}
            fixed_cols  = {}
            cur_mes     = None

            for ci, v in enumerate(vals):
                vu = v.upper().replace(" ", "").replace("º","").replace("º","")
                vl = v.lower()

                if not v:
                    continue

                # Columnas fijas
                if "PREST" in vu:
                    fixed_cols["nprsto"] = ci
                elif vu in ("NFRA", "NFR", "NOFRA"):
                    fixed_cols["nfra"] = ci
                elif "CLIENT" in vu:
                    fixed_cols["cliente"] = ci
                elif vu in ("OBRA", "NOOBRA") or (vu.startswith("OBRA") and len(vu) < 8):
                    fixed_cols["obra"] = ci
                elif vu in ("NOBRE", "NOOBRA", "NOBRA"):
                    fixed_cols["nobre"] = ci

                # Nombre del mes
                for i, m in enumerate(MESES):
                    if m in vl:
                        cur_mes = MESES_ES[i]
                        print(f"  Mes '{cur_mes}' detectado en fila {ri+1}, col {ci+1}")
                        break

                # Número de día
                try:
                    d = int(float(v))
                    if 1 <= d <= 31:
                        day_col_map[ci] = d
                except (ValueError, TypeError):
                    pass

            print(f"  Cabecera fila {ri+1} — fixed: {fixed_cols} — días: {len(day_col_map)} cols")
            continue

        # ── Fila de datos
        if cur_mes is None or not day_col_map:
            continue
        if "obra" not in fixed_cols and "cliente" not in fixed_cols:
            continue

        obra    = vals[fixed_cols["obra"]]    if "obra"    in fixed_cols else ""
        cliente = vals[fixed_cols["cliente"]] if "cliente" in fixed_cols else ""
        if not obra and not cliente:
            continue

        nprsto = vals[fixed_cols["nprsto"]] if "nprsto" in fixed_cols else ""
        nfra   = vals[fixed_cols["nfra"]]   if "nfra"   in fixed_cols else ""

        dias_activos = 0
        for ci, day_num in day_col_map.items():
            cell = row[ci]
            rgb  = color_rgb(cell)

            active  = False
            weekend = False

            if rgb:
                if is_red(rgb):
                    active = True; weekend = True
                elif is_yellow(rgb):
                    active = True
                else:
                    # Cualquier relleno no blanco/negro se considera activo
                    active = True

            # Fallback: celda con valor
            if not active and cell.value is not None and str(cell.value).strip():
                active = True

            if active:
                key = str(day_num)
                if key not in result[cur_mes]:
                    result[cur_mes][key] = []
                result[cur_mes][key].append({
                    "nprsto":   nprsto,
                    "nfra":     nfra,
                    "cliente":  cliente,
                    "obra":     obra,
                    "weekend":  weekend,
                })
                dias_activos += 1

        if dias_activos:
            print(f"    Obra '{obra[:40]}' → {dias_activos} días activos")

    # Resumen
    print("\n=== RESUMEN ===")
    for m in MESES_ES:
        n = len(result[m])
        total = sum(len(v) for v in result[m].values())
        print(f"  {m}: {n} días, {total} entradas de obra")

    return result

def main():
    if len(sys.argv) < 2:
        # Buscar Excel en la misma carpeta
        script_dir = Path(__file__).parent
        excels = list(script_dir.glob("*.xlsx")) + list(script_dir.glob("*.xls")) + list(script_dir.glob("*.xlsm"))
        if not excels:
            print("Uso: python3 generar_datos.py archivo.xlsx")
            print("O bien coloca el Excel en la misma carpeta que este script.")
            sys.exit(1)
        filepath = str(excels[0])
        print(f"Excel encontrado: {filepath}")
    else:
        filepath = sys.argv[1]

    if not Path(filepath).exists():
        print(f"Error: no se encuentra el archivo '{filepath}'")
        sys.exit(1)

    data = parse_excel(filepath)

    out = Path(__file__).parent / "data.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Datos guardados en: {out}")
    print("Ahora abre index.html en el navegador.")

if __name__ == "__main__":
    main()
